import openpyxl
import pygame
import sys
import random
import mysql.connector
import string

# Inicialización de pygame
pygame.init()

# Configuración de conexión a la base de datos
def obtener_datos_infracciones():
    conexion = mysql.connector.connect(
        host="localhost",
        user="root",
        password="",
        database="infracciones_db"
    )
    cursor = conexion.cursor()
    cursor.execute("SELECT vehiculo, color, placa, costo_multa, descripcion FROM infracciones")
    datos = cursor.fetchall()
    conexion.close()
    return datos

def insertar_infraccion(vehiculo, color, placa, costo_multa, descripcion):
    conexion = mysql.connector.connect(
        host="localhost",
        user="root",
        password="",
        database="infracciones_db"
    )
    cursor = conexion.cursor()
    sql = "INSERT INTO infracciones (vehiculo, color, placa, costo_multa, descripcion) VALUES (%s, %s, %s, %s, %s)"
    valores = (vehiculo, color, placa, costo_multa, descripcion)
    cursor.execute(sql, valores)
    conexion.commit()
    conexion.close()

# Función para generar el documento de Excel
def crear_documento_excel(datos):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Infracciones"
    
    encabezados = ["Vehículo", "Color", "Placa", "Costo Multa", "Descripción"]
    for col, encabezado in enumerate(encabezados, start=1):
        sheet.cell(row=1, column=col, value=encabezado)
    
    for row, dato in enumerate(datos, start=2):
        for col, valor in enumerate(dato, start=1):
            sheet.cell(row=row, column=col, value=valor)
    
    workbook.save("infracciones.xlsx")
    print("Documento Excel generado como 'infracciones.xlsx'")

# Tamaño de la ventana
WINDOW_WIDTH, WINDOW_HEIGHT = 900, 800
window = pygame.display.set_mode((WINDOW_WIDTH, WINDOW_HEIGHT))
pygame.display.set_caption("Simulador de Transito con Semáforos")

# Colores
RED = (255, 0, 0)
GREEN = (0, 255, 0)
BLUE = (0, 0, 255)
WHITE = (255, 255, 255)
BLACK = (0, 0, 0)
YELLOW = (255, 255, 0)
ALERT_BG_COLOR = (255, 255, 153)
GRAY = (169, 169, 169)  # Color de fondo gris


# Variable para controlar la actualización de la tabla
tabla_actualizada = True

# Variables para la alerta
alert_active = False
alert_message = ""
alert_start_time = 0
alert_duration = 3000

# Cargar la imagen de fondo
background_image = pygame.image.load('mapa.png')
ORIGINAL_WIDTH, ORIGINAL_HEIGHT = background_image.get_size()
game_surface = pygame.Surface((ORIGINAL_WIDTH, ORIGINAL_HEIGHT))

# Cargar imágenes del semáforo y autos
semaforo_rojo_img = pygame.transform.scale(pygame.image.load('colores/rojo.png'), (30, 60))
semaforo_verde_img = pygame.transform.scale(pygame.image.load('colores/verde.png'), (30, 60))
semaforo_amarillo_img = pygame.transform.scale(pygame.image.load('colores/amarillo.png'), (30, 60))

car_red_img = pygame.transform.scale(pygame.image.load('autos/car.png'), (30, 15))
car_blue_img = pygame.transform.rotate(pygame.transform.scale(pygame.image.load('autos/bus.png'), (30, 15)), 90)

# Función para mostrar una alerta temporal
def mostrar_alerta(vehiculo, tipo_infraccion):
    global alert_active, alert_message, alert_start_time
    alert_active = True
    alert_message = f"Infracción: {vehiculo} - {tipo_infraccion}"
    alert_start_time = pygame.time.get_ticks()

def dibujar_alerta(surface):
    global alert_active
    if alert_active:
        elapsed_time = pygame.time.get_ticks() - alert_start_time
        if elapsed_time < alert_duration:
            font = pygame.font.SysFont(None, 30, bold=True)
            text = font.render(alert_message, True, BLACK)
            text_rect = text.get_rect(center=(WINDOW_WIDTH // 2, 100))
            pygame.draw.rect(surface, ALERT_BG_COLOR, text_rect.inflate(20, 20))
            surface.blit(text, text_rect)
        else:
            alert_active = False

# Función para generar una placa aleatoria
def generar_placa():
    letras = ''.join(random.choices(string.ascii_uppercase, k=4))
    numeros = ''.join(random.choices(string.digits, k=3))
    return letras + numeros

# Configuración de la tabla
table_x = 150
table_y = WINDOW_HEIGHT - 200
table_height = 150
num_rows = 3
cell_height = table_height // num_rows
column_widths = [40, 80, 80, 100, 70, 180]
headers = ["N°", "Tipo", "Color", "Placa", "Multa $", "Descripción"]

# Coordenadas para los botones de color
color_buttons_x = 50
color_buttons_y = table_y - 110
red_button_rect = pygame.Rect(color_buttons_x - 15, color_buttons_y - 15, 30, 30)
green_button_rect = pygame.Rect(color_buttons_x - 15, color_buttons_y + 15, 30, 30)
blue_button_rect = pygame.Rect(color_buttons_x - 15, color_buttons_y + 45, 30, 30)

# Control de paginación y pausa
offset = 0
rows_per_page = 3
paused = False

# Función para dibujar los encabezados de la tabla
def draw_table_headers(surface):
    font = pygame.font.SysFont(None, 24, bold=True)
    x_offset = table_x
    for i, header in enumerate(headers):
        text = font.render(header, True, BLACK)
        surface.blit(text, (x_offset + 15, table_y - cell_height + 25))
        x_offset += column_widths[i]

# Función para dibujar los botones de colores
def draw_color_buttons(surface):
    font = pygame.font.SysFont(None, 24, bold=True)
    pygame.draw.circle(surface, RED, red_button_rect.center, 15)
    surface.blit(font.render("Pausar", True, BLACK), (color_buttons_x + 25, color_buttons_y - 10))
    pygame.draw.circle(surface, GREEN, green_button_rect.center, 15)
    surface.blit(font.render("Reanudar", True, BLACK), (color_buttons_x + 25, color_buttons_y + 20))
    pygame.draw.circle(surface, BLUE, blue_button_rect.center, 15)
    surface.blit(font.render("Generar Excel", True, BLACK), (color_buttons_x + 25, color_buttons_y + 50))

# Función para dibujar la tabla con los datos
def draw_table_with_data(surface, datos, offset):
    font = pygame.font.SysFont(None, 24)
    row_offset = 0
    x_offset = table_x
    for width in column_widths:
        pygame.draw.line(surface, BLACK, (x_offset, table_y), (x_offset, table_y + table_height), 2)
        x_offset += width
    pygame.draw.line(surface, BLACK, (table_x + sum(column_widths), table_y), (table_x + sum(column_widths), table_y + table_height), 2)
    for row in range(num_rows + 1):
        pygame.draw.line(surface, BLACK, (table_x, table_y + row * cell_height), (table_x + sum(column_widths), table_y + row * cell_height), 2)
    for i in range(offset, min(offset + num_rows, len(datos))):
        vehiculo, color, placa, costo_multa, descripcion = datos[i]
        data = [i + 1 + offset, vehiculo, color, placa, f"{costo_multa:.2f}", descripcion]
        x_offset = table_x
        for j, item in enumerate(data):
            surface.blit(font.render(str(item), True, BLACK), (x_offset + 10, table_y + row_offset * cell_height + 5))
            x_offset += column_widths[j]
        row_offset += 1

# Dibujar flechas de navegación
def draw_navigation_arrows(surface):
    pygame.draw.polygon(surface, BLUE, [
        (table_x + 10, table_y + table_height + 20), 
        (table_x + 30, table_y + table_height + 10), 
        (table_x + 30, table_y + table_height + 30)
    ])
    pygame.draw.polygon(surface, BLUE, [
        (table_x + sum(column_widths) - 10, table_y + table_height + 20), 
        (table_x + sum(column_widths) - 30, table_y + table_height + 10), 
        (table_x + sum(column_widths) - 30, table_y + table_height + 30)
    ])

def handle_arrow_click(mouse_pos):
    global offset
    if table_x + 10 < mouse_pos[0] < table_x + 30 and table_y + table_height + 10 < mouse_pos[1] < table_y + table_height + 30:
        if offset > 0:
            offset -= rows_per_page
    elif table_x + sum(column_widths) - 30 < mouse_pos[0] < table_x + sum(column_widths) - 10 and table_y + table_height + 10 < mouse_pos[1] < table_y + table_height + 30:
        if offset + rows_per_page < len(datos_infracciones):
            offset += rows_per_page

# Clase Semaforo
class Semaforo:
    def __init__(self, x, y, initial_color, duration_green=5000, duration_yellow=2000, duration_red=3000):
        self.x, self.y = x, y
        self.color = initial_color
        self.durations = {GREEN: duration_green, YELLOW: duration_yellow, RED: duration_red}
        self.last_switch = pygame.time.get_ticks()

    def update(self):
        elapsed_time = pygame.time.get_ticks() - self.last_switch
        if elapsed_time >= self.durations[self.color]:
            self.color = {GREEN: YELLOW, YELLOW: RED, RED: GREEN}[self.color]
            self.last_switch = pygame.time.get_ticks()

    def draw(self, surface):
        img = {RED: semaforo_rojo_img, YELLOW: semaforo_amarillo_img, GREEN: semaforo_verde_img}[self.color]
        surface.blit(img, (self.x, self.y))

# Clase Carro modificada para almacenar infracciones y generar placas
class Car:
    def __init__(self, image, x, y, direction, speed, free_zone_limit, pass_zone_limit, name, color, infringement_interval):
        self.image = image
        self.x, self.y = x, y
        self.direction = direction
        self.free_zone_limit = free_zone_limit
        self.pass_zone_limit = pass_zone_limit
        self.speed = self.original_speed = speed
        self.name = name
        self.color = color
        self.placa = generar_placa()
        self.speeding = False
        self.in_free_zone = True
        self.infringement = None
        self.last_infringement_time = 0  # Para almacenar el tiempo de la última infracción
        self.infringement_interval = infringement_interval  # Intervalo de tiempo para permitir la siguiente infracción
        self.ignore_lights = False

    def move(self, semaforo):
        if paused:
            return

        current_time = pygame.time.get_ticks()
        
        # Chequeo de zona libre y desplazamiento
        if self.in_free_zone:
            if (self.direction == 'horizontal' and self.x >= self.free_zone_limit) or \
               (self.direction == 'vertical' and self.y >= self.free_zone_limit):
                self.in_free_zone = False
            else:
                self._move_forward()
        else:
            # Chequeo de luz roja y tiempo de la última infracción
            if not self.ignore_lights:
                if ((self.direction == 'horizontal' and self.x < self.pass_zone_limit) or
                    (self.direction == 'vertical' and self.y < self.pass_zone_limit)):
                    
                    # Verificar infracción por ignorar luz roja
                    if semaforo.color == RED:
                        if (current_time - self.last_infringement_time) >= self.infringement_interval:
                            if random.choice([True, False]):
                                self.infringement = "Ignoró el paso peatonal"
                                self.registrar_infraccion(20)
                                self.last_infringement_time = current_time  # Actualizar tiempo de infracción
                                self.ignore_lights = True
                                mostrar_alerta(self.name, self.infringement)
                        else:
                            return
                    else:
                        self._move_forward()
                else:
                    self.check_speeding(current_time)
                    self._move_forward()
            else:
                self._move_forward()

    def check_speeding(self, current_time):
        # Verificar infracción por exceso de velocidad
        if not self.speeding and (current_time - self.last_infringement_time) >= self.infringement_interval:
            if random.choice([True, False]):
                self.speed = self.original_speed * 10
                self.speeding = True
                self.infringement = "Exceso de velocidad"
                self.registrar_infraccion(70)
                self.last_infringement_time = current_time  # Actualizar tiempo de infracción
                mostrar_alerta(self.name, self.infringement)

    def registrar_infraccion(self, costo_multa):
        global tabla_actualizada
        tipo = "Automóvil" if self.name == "Auto Rojo" else "Bus"
        descripcion = self.infringement
        insertar_infraccion(tipo, self.color, self.placa, costo_multa, descripcion)
        tabla_actualizada = True

    # Método para avanzar en la dirección del auto
    def _move_forward(self):
        if self.direction == 'horizontal':
            self.x += self.speed
            if self.x > ORIGINAL_WIDTH:
                self._reset_position()
        elif self.direction == 'vertical':
            self.y += self.speed
            if self.y > ORIGINAL_HEIGHT:
                self._reset_position()

    def _reset_position(self):
        if self.direction == 'horizontal':
            self.x = -30
        else:
            self.y = -15
        self.in_free_zone = True
        self.ignore_lights = False
        self.speed = self.original_speed
        self.speeding = False
        self.placa = generar_placa()

    def draw(self, surface):
        surface.blit(self.image, (self.x, self.y))

# Crear autos y semáforos
car_red = Car(car_red_img, 60, 285, 'horizontal', 2, 130, 140, "Auto Rojo", "Rojo", 30000)  # 15 segundos
car_blue = Car(car_blue_img, 265, 50, 'vertical', 3, 160, 180, "Bus Azul", "Azul", 60000)  # 27 segundos
semaforo_horizontal = Semaforo(250, 285, initial_color=GREEN, duration_green=10000, duration_yellow=3000, duration_red=7000)
semaforo_vertical = Semaforo(265, 200, initial_color=RED, duration_green=10000, duration_yellow=3000, duration_red=7000)

# Cargar los datos de la base de datos
datos_infracciones = obtener_datos_infracciones()

# Bucle principal del juego
clock = pygame.time.Clock()
running = True
while running:
    window.fill(GRAY)
    game_surface.blit(background_image, (0, 0))

    # Actualizar datos de la tabla si hay cambios
    if tabla_actualizada:
        datos_infracciones = obtener_datos_infracciones()
        tabla_actualizada = False

    # Detectar posición del ratón para botones de color
    mouse_x, mouse_y = pygame.mouse.get_pos()
    if red_button_rect.collidepoint(mouse_x, mouse_y) or green_button_rect.collidepoint(mouse_x, mouse_y) or blue_button_rect.collidepoint(mouse_x, mouse_y):
        pygame.mouse.set_cursor(pygame.SYSTEM_CURSOR_HAND)
    else:
        pygame.mouse.set_cursor(pygame.SYSTEM_CURSOR_ARROW)

    for event in pygame.event.get():
        if event.type == pygame.QUIT:
            running = False
        elif event.type == pygame.MOUSEBUTTONDOWN:
            if red_button_rect.collidepoint(event.pos):
                paused = True
            elif green_button_rect.collidepoint(event.pos):
                paused = False
            elif blue_button_rect.collidepoint(event.pos):
                crear_documento_excel(datos_infracciones)
            elif table_x + 10 < event.pos[0] < table_x + 30 or table_x + sum(column_widths) - 30 < event.pos[0] < table_x + sum(column_widths) - 10:
                handle_arrow_click(event.pos)

    if not paused:
        semaforo_horizontal.update()
        semaforo_vertical.update()
        car_red.move(semaforo_horizontal)
        car_blue.move(semaforo_vertical)

    semaforo_horizontal.draw(game_surface)
    semaforo_vertical.draw(game_surface)
    car_red.draw(game_surface)
    car_blue.draw(game_surface)

    # Dibujar elementos de la UI y tabla
    window.blit(game_surface, ((WINDOW_WIDTH - ORIGINAL_WIDTH) // 2, ((WINDOW_HEIGHT - ORIGINAL_HEIGHT) // 2) - 150))
    draw_color_buttons(window)
    draw_table_headers(window)
    draw_table_with_data(window, datos_infracciones, offset)
    draw_navigation_arrows(window)
    dibujar_alerta(window)

    pygame.display.flip()
    clock.tick(30)

pygame.quit()
sys.exit()
